import PySimpleGUI as sg
import easygui
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import win32com.client as win32
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
import tempfile
import os


sg.theme('DarkBlue9')
# This design pattern simulates button callbacks
# This implementation uses a simple "Dispatch Dictionary" to store events and functions

# The callback functions
version = __version__ = "1.2.4"

def button1():
    path = easygui.fileopenbox(default="HOME/Downloads/*.xlsx", filetypes='*.xlsx')
    try:
        workbook = openpyxl.load_workbook(path)
    except Exception as e:
        try:
            workbook = openpyxl.load_workbook(path)
        except Exception as e:
            sg.popup_auto_close('не выбран файл!')
            sg.popup_ok(e)
            workbook = None
        return (workbook)
    #path = easygui.fileopenbox(default= '*.xlsx', filetypes= '*.xlsx')
    path2 = easygui.filesavebox(default="HOME/APPDATA/RPBeta/Egoryevsk.xlsx", filetypes='*.xslx')
    #workbook = openpyxl.load_workbook(path)
    worksheet = workbook["Лист 1"]
    filetypes = ['*.xlsx', "Excel"]
    default = '*'
    worksheet.unmerge_cells('A1:P1')
    worksheet.unmerge_cells('A2:P2')
    worksheet.unmerge_cells('A3:P3')
    worksheet.unmerge_cells('A4:P4')
    worksheet.unmerge_cells('A5:P5')
    worksheet['D4'] = '= SUBTOTAL(9,D10:D119)'
    worksheet['E4'] = '= SUBTOTAL(9,E10:E119)'
    worksheet['G4'] = '= SUBTOTAL(9,G10:G119)'
    worksheet['F4'].number_format = FORMAT_PERCENTAGE_00
    worksheet['H4'].number_format = FORMAT_PERCENTAGE_00
    worksheet['F4'] = '= E4/D4'
    worksheet['H4'] = '= G4/D4'
    worksheet['A4'] = 'ИТОГИ'
    worksheet['G1'] = 'RNISka Reports 1.2.4'
    worksheet['G1'].font = Font(name='Tahoma', size=9, color="FF0000", italic=True)
    worksheet.merge_cells('D4:D5')
    worksheet.merge_cells('E4:E5')
    worksheet.merge_cells('G4:G5')
    worksheet.merge_cells('F4:F5')
    worksheet.merge_cells('H4:H5')
    worksheet['D4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['E4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['G4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['F4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['H4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet.merge_cells('A4:C5')
    worksheet['A4'].alignment = Alignment(horizontal="left", vertical="center")
    worksheet['D4'].font = Font(bold=True, size=12)
    worksheet['E4'].font = Font(bold=True, size=12)
    worksheet['G4'].font = Font(bold=True, size=12)
    worksheet['F4'].font = Font(bold=True, size=12)
    worksheet['H4'].font = Font(bold=True, size=12)
    worksheet.delete_cols(9, 17)
    ws1 = worksheet.title = "Егорьевск"
    FullRange = "C6:" + get_column_letter(worksheet.max_column) \
    + str(worksheet.max_row)
    worksheet.auto_filter.ref = FullRange
    ws1 = worksheet.auto_filter.add_filter_column(0, ["167", "168", "169", "171", "172", "173", "174", "175", "176", \
                                                      "177", "178", "180", "181", "182", "183", "184", "185", "186",\
                                                      "187", "188", "189", "190", "193", "194", "678", "1853", "1855",\
                                                      "2102", "2600", "3252", "3253", "3254", "179*"])
    workbook.save(path2)
    sg.popup_ok('Файл готов!')

def button2():
    path = easygui.fileopenbox(default="HOME/Downloads/*.xlsx", filetypes='*.xlsx')
    try:
        workbook = openpyxl.load_workbook(path)
    except Exception as e:
        try:
            workbook = openpyxl.load_workbook(path)
        except Exception as e:
            sg.popup_auto_close('не выбран файл!')
            sg.popup_ok(e)
            workbook = None
        return (workbook)
    #path = easygui.fileopenbox(default= '*.xlsx', filetypes= '*.xlsx')
    path2 = easygui.filesavebox(default="HOME/APPDATA/RPBeta/Ramenskoye.xlsx", filetypes= '*.xslx')
    #workbook = openpyxl.load_workbook(path)
    worksheet = workbook["Лист 1"]
    filetypes=['*.xlsx', "Excel"]
    default='*'
    worksheet = workbook["Лист 1"]
    worksheet.unmerge_cells('A1:P1')
    worksheet.unmerge_cells('A2:P2')
    worksheet.unmerge_cells('A3:P3')
    worksheet.unmerge_cells('A4:P4')
    worksheet.unmerge_cells('A5:P5')
    worksheet['D4'] = '= SUBTOTAL(9,D9:D125)'
    worksheet['E4'] = '= SUBTOTAL(9,E9:E125)'
    worksheet['G4'] = '= SUBTOTAL(9,G9:G125)'
    worksheet['F4'].number_format = FORMAT_PERCENTAGE_00
    worksheet['H4'].number_format = FORMAT_PERCENTAGE_00
    worksheet['F4'] = '= E4/D4'
    worksheet['H4'] = '= G4/D4'
    worksheet['A4'] = 'ИТОГИ'
    worksheet['G1'] = 'RNISka Reports 1.2.4'
    worksheet['G1'].font = Font(name='Tahoma', size=9, color="FF0000", italic=True)
    worksheet.merge_cells('D4:D5')
    worksheet.merge_cells('E4:E5')
    worksheet.merge_cells('G4:G5')
    worksheet.merge_cells('F4:F5')
    worksheet.merge_cells('H4:H5')
    worksheet['D4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['E4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['G4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['F4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['H4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet.merge_cells('A4:C5')
    worksheet['A4'].alignment = Alignment(horizontal="left", vertical="center")
    worksheet['D4'].font = Font(bold=True, size=12)
    worksheet['E4'].font = Font(bold=True, size=12)
    worksheet['G4'].font = Font(bold=True, size=12)
    worksheet['F4'].font = Font(bold=True, size=12)
    worksheet['H4'].font = Font(bold=True, size=12)
    worksheet.delete_cols(9, 17)
    ws1 = worksheet.title = "Раменское"
    FullRange = "C6:" + get_column_letter(worksheet.max_column) \
    + str(worksheet.max_row)
    worksheet.auto_filter.ref = FullRange
    ws1 = worksheet.auto_filter.add_filter_column(0, ["1122", "1123", "1124", "1125", "1126", "1127", "1128", "1129", \
                                                      "1130", "1131", "1132", "1133", "1134", "1135", "1136", "1137", \
                                                      "1138", "1139", "1140", "1141", "1142", "1147", "1153", "1154", \
                                                      "1156", "1161", "1689", "1723", "1724", "1769", "1770", "1771", \
                                                      "2104", "2120", "2185", "2811", "3063", "3066", "3067", "3074"])
    workbook.save(path2)
    sg.popup_ok('Файл готов!')

def button3():
    path = easygui.fileopenbox(default="HOME/Downloads/*.xlsx", filetypes='*.xlsx')
    try:
        workbook = openpyxl.load_workbook(path)
    except Exception as e:
        try:
            workbook = openpyxl.load_workbook(path)
        except Exception as e:
            sg.popup_auto_close('не выбран файл!')
            sg.popup_ok(e)
            workbook = None
        return (workbook)
#   path = easygui.fileopenbox(default= '*.xlsx', filetypes= '*.xlsx')
    path2 = easygui.filesavebox(default="HOME/APPDATA/RPBeta/Shatoora.xlsx", filetypes= '*.xslx')
#   workbook = openpyxl.load_workbook(path)
    worksheet = workbook["Лист 1"]
    filetypes=['*.xlsx', "Excel"]
    default='*'
    worksheet = workbook["Лист 1"]
    worksheet.unmerge_cells('A1:P1')
    worksheet.unmerge_cells('A2:P2')
    worksheet.unmerge_cells('A3:P3')
    worksheet.unmerge_cells('A4:P4')
    worksheet.unmerge_cells('A5:P5')
    worksheet['D4'] = '= SUBTOTAL(9,D8:D127)'
    worksheet['E4'] = '= SUBTOTAL(9,E8:E127)'
    worksheet['G4'] = '= SUBTOTAL(9,G8:G127)'
    worksheet['F4'].number_format = FORMAT_PERCENTAGE_00
    worksheet['H4'].number_format = FORMAT_PERCENTAGE_00
    worksheet['F4'] = '= E4/D4'
    worksheet['H4'] = '= G4/D4'
    worksheet['A4'] = 'ИТОГИ'
    worksheet['G1'] = 'RNISka Reports 1.2.4'
    worksheet['G1'].font = Font(name='Tahoma', size=9, color="FF0000", italic=True)
    worksheet.merge_cells('D4:D5')
    worksheet.merge_cells('E4:E5')
    worksheet.merge_cells('G4:G5')
    worksheet.merge_cells('F4:F5')
    worksheet.merge_cells('H4:H5')
    worksheet['D4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['E4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['G4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['F4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['H4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet.merge_cells('A4:C5')
    worksheet['A4'].alignment = Alignment(horizontal="left", vertical="center")
    worksheet['D4'].font = Font(bold=True, size=12)
    worksheet['E4'].font = Font(bold=True, size=12)
    worksheet['G4'].font = Font(bold=True, size=12)
    worksheet['F4'].font = Font(bold=True, size=12)
    worksheet['H4'].font = Font(bold=True, size=12)
    worksheet.delete_cols(9, 17)
    ws1 = worksheet.title = "Шатура"
    FullRange = "C6:" + get_column_letter(worksheet.max_column) \
    + str(worksheet.max_row)
    worksheet.auto_filter.ref = FullRange
    ws1 = worksheet.auto_filter.add_filter_column(0, ["1513", "1172", "1993", "1514", "1515", "1516", "1988", "1519", \
                                                      "1697", "3234", "2082", "1517", "1518", "1521", "1543", "1522", \
                                                      "2491", "1523", "1524", "1525", "1544", "2609", "1546", "1547", \
                                                      "1526", "1527", "1528", "1529", "1530", "1548", "1531", "1532", \
                                                      "1550", "2565", "1549", "1545", "1991", "1896", "1534", "1538", \
                                                      "1539", "1811", "1540", "1541", "1812", "1542", "1759"])
    workbook.save(path2)
    sg.popup_ok('Файл готов!')

def button4():
    path = easygui.fileopenbox(default="HOME/Downloads/*.xlsx", filetypes='*.xlsx')
    try:
        workbook = openpyxl.load_workbook(path)
    except Exception as e:
        try:
            workbook = openpyxl.load_workbook(path)
        except Exception as e:
            sg.popup_auto_close('не выбран файл!')
            sg.popup_ok(e)
            workbook = None
        return (workbook)
#   path = easygui.fileopenbox(default= '*.xlsx', filetypes= '*.xlsx')
    path2 = easygui.filesavebox(default="HOME/Downloads/map4.xlsx", filetypes= '*.xslx')
#   workbook = openpyxl.load_workbook(path)
    worksheet = workbook["Лист 1"]
    filetypes=['*.xlsx', "Excel"]
    default='*'
    worksheet = workbook["Лист 1"]
    worksheet.unmerge_cells('A1:P1')
    worksheet.unmerge_cells('A2:P2')
    worksheet.unmerge_cells('A3:P3')
    worksheet.unmerge_cells('A4:P4')
    worksheet.unmerge_cells('A5:P5')
    worksheet['D4'] = '= SUBTOTAL(9,D10:D119)'
    worksheet['E4'] = '= SUBTOTAL(9,E10:E119)'
    worksheet['G4'] = '= SUBTOTAL(9,G10:G119)'
    worksheet['F4'].number_format = FORMAT_PERCENTAGE_00
    worksheet['H4'].number_format = FORMAT_PERCENTAGE_00
    worksheet['F4'] = '= E4/D4'
    worksheet['H4'] = '= G4/D4'
    worksheet['A4'] = 'ИТОГИ'
    worksheet['G1'] = 'RNISka Reports 1.2.4'
    worksheet['G1'].font = Font(name='Tahoma', size=9, color="FF0000", italic=True)
    worksheet.merge_cells('D4:D5')
    worksheet.merge_cells('E4:E5')
    worksheet.merge_cells('G4:G5')
    worksheet.merge_cells('F4:F5')
    worksheet.merge_cells('H4:H5')
    worksheet['D4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['E4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['G4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['F4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['H4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet.merge_cells('A4:C5')
    worksheet['A4'].alignment = Alignment(horizontal="left", vertical="center")
    worksheet['D4'].font = Font(bold=True, size=12)
    worksheet['E4'].font = Font(bold=True, size=12)
    worksheet['G4'].font = Font(bold=True, size=12)
    worksheet['F4'].font = Font(bold=True, size=12)
    worksheet['H4'].font = Font(bold=True, size=12)
    worksheet.delete_cols(9, 17)
    ws1 = worksheet.title = "Егорьевск"
    FullRange = "C6:" + get_column_letter(worksheet.max_column) \
    + str(worksheet.max_row)
    worksheet.auto_filter.ref = FullRange
    ws1 = worksheet['D4'] = '= SUBTOTAL(9,D10:D119)'
    ws1 = worksheet['E4'] = '= SUBTOTAL(9,E10:E119)'
    ws1 = worksheet['G4'] = '= SUBTOTAL(9,G10:G119)'
    ws1 = worksheet.auto_filter.add_filter_column(0, ["167", "168", "169", "171", "172", "173", "174", "175", "176", \
                                                      "177", "178", "180", "181", "182", "183", "184", "185", "186", \
                                                      "187", "188", "189", "190", "193", "194", "678", "1853", "1855", \
                                                      "2102", "2600", "3252", "3253", "3254", "179*"])

    source = workbook.active
    target = workbook.copy_worksheet(source)
    worksheet = workbook["Егорьевск Copy"]
    ws2 = worksheet.title = "Раменское"
    FullRange = "C6:" + get_column_letter(worksheet.max_column) \
    + str(worksheet.max_row)
    worksheet.auto_filter.ref = FullRange
    ws2 = worksheet['D4'] = '= SUBTOTAL(9,D9:D125)'
    ws2 = worksheet['E4'] = '= SUBTOTAL(9,E9:E125)'
    ws2 = worksheet['G4'] = '= SUBTOTAL(9,G9:G125)'
    ws2 = worksheet.auto_filter.add_filter_column(0, ["1122", "1123", "1124", "1125", "1126", "1127", "1128", "1129", \
                                                      "1130", "1131", "1132", "1133", "1134", "1135", "1136", "1137", \
                                                      "1138", "1139", "1140", "1141", "1142", "1147", "1153", "1154", \
                                                      "1156", "1161", "1689", "1723", "1724", "1769", "1770", "1771", \
                                                      "2104", "2120", "2185", "2811", "3063", "3066", "3067", "3074"])

    source = workbook.active
    target = workbook.copy_worksheet(source)
    worksheet = workbook["Егорьевск Copy"]
    ws3 = worksheet.title = "Шатура"
    FullRange = "C6:" + get_column_letter(worksheet.max_column) \
    + str(worksheet.max_row)
    worksheet.auto_filter.ref = FullRange
    ws3 = worksheet['D4'] = '= SUBTOTAL(9,D8:D127)'
    ws3 = worksheet['E4'] = '= SUBTOTAL(9,E8:E127)'
    ws3 = worksheet['G4'] = '= SUBTOTAL(9,G8:G127)'
    ws3 = worksheet.auto_filter.add_filter_column(0, ["1513", "1172", "1993", "1514", "1515", "1516", "1988", "1519", \
                                                      "1697", "3234", "2082", "1517", "1518", "1521", "1543", "1522", \
                                                      "2491", "1523", "1524", "1525", "1544", "2609", "1546", "1547", \
                                                      "1526", "1527", "1528", "1529", "1530", "1548", "1531", "1532", \
                                                      "1550", "2565", "1549", "1545", "1991", "1896", "1534", "1538", \
                                                      "1539", "1811", "1540", "1541", "1812", "1542", "1759"])
    workbook.save(path2)
    sg.popup_ok('Файл готов!')


def button5():
    # рабочая кнопка для конвертации
    size = (50, 3)
    auto_size_button = True
    path = easygui.fileopenbox(default="HOME/Downloads/*.xls", filetypes='*.xls')
    fname = path
    excel = win32.gencache.EnsureDispatch('Excel.Application')
#    temp = tempfile.TemporaryFile()
#    sg.popup_ok('начинается проверка условий')
    try:
        wb = excel.Workbooks.Open(fname)
        wb.SaveAs(fname + "x", FileFormat=51)  # FileFormat = 51 is for .xlsx extension
        wb.Close()
        excel.Application.Quit()
        sg.popup_ok('сконвертировано')
#       temp.close()
    except Exception as e:
        try:
            wb = excel.Workbooks.Open(fname)
        except Exception as e:
            sg.popup_ok(e)
            sg.popup_ok('неудача')
            wb = None
            return (wb)
    finally:
        sg.popup_ok('завершение') # задача при которой условия частично выполнены.


def button6():
    path = easygui.fileopenbox(default="HOME/Downloads/*.xlsx", filetypes='*.xlsx')
    try:
        workbook = openpyxl.load_workbook(path)
    except Exception as e:
        try:
            workbook = openpyxl.load_workbook(path)
            openpyxl.Visible = True
        except Exception as e:
            sg.popup_auto_close('не выбран файл!')
            sg.popup_ok(e)
            workbook = None
        return (workbook)

#   workbook = openpyxl.load_workbook(path)
    worksheet = workbook.active
    filetypes = ['*.xlsx', "Excel"]
    default = '*'
    worksheet.unmerge_cells('A1:P1')
    worksheet.unmerge_cells('A2:P2')
    worksheet.unmerge_cells('A3:P3')
    worksheet.unmerge_cells('A4:P4')
    worksheet.unmerge_cells('A5:P5')
    worksheet['D4'] = '= SUBTOTAL(9,D10:D125)'
    worksheet['E4'] = '= SUBTOTAL(9,E10:E125)'
    worksheet['G4'] = '= SUBTOTAL(9,G10:G125)'
    worksheet['F4'].number_format = FORMAT_PERCENTAGE_00
    worksheet['H4'].number_format = FORMAT_PERCENTAGE_00
    worksheet['F4'] = '= E4/D4'
    worksheet['H4'] = '= G4/D4'
    worksheet['A4'] = 'ИТОГИ'
    worksheet['G1'] = 'RNISka Reports 1.2.4'
    worksheet['G1'].font = Font(name='Tahoma', size=9, color="FF0000", italic=True)
    worksheet.merge_cells('D4:D5')
    worksheet.merge_cells('E4:E5')
    worksheet.merge_cells('G4:G5')
    worksheet.merge_cells('F4:F5')
    worksheet.merge_cells('H4:H5')
    worksheet['D4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['E4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['G4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['F4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['H4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet.merge_cells('A4:C5')
    worksheet['A4'].alignment = Alignment(horizontal="left", vertical="center")
    worksheet['D4'].font = Font(bold=True, size=12)
    worksheet['E4'].font = Font(bold=True, size=12)
    worksheet['G4'].font = Font(bold=True, size=12)
    worksheet['F4'].font = Font(bold=True, size=12)
    worksheet['H4'].font = Font(bold=True, size=12)

    worksheet.delete_cols(9, 17)
    ws1 = worksheet.title = "сходы"
    FullRange = "C6:" + get_column_letter(worksheet.max_column) \
            + str(worksheet.max_row)
    worksheet.auto_filter.ref = FullRange
    ws1 = worksheet.auto_filter.add_filter_column(0, ["173", "177", "179*", "193", "182", "186", "1122", "1123",    \
                                                          "2104", "1124", "2811", "1769", "1125", "1131", "1132",   \
                                                          "1771", "1127", "1128", "1129", "1130", "1153", "1156",   \
                                                          "1134", "1770", "1136", "1137", "1138", "1139", "1141",   \
                                                          "1723", "1689", "2185", "1126", "1161", "1172", "1988",   \
                                                          "1519", "1697", "3234", "1517", "1518", "1543", "1522",   \
                                                          "1524", "1525", "1546", "1526", "1528", "1532"])
#    excel.Workbooks.Open(path)
    workbook.save(path)
    sg.popup_ok('Файл готов!')


def button7(*args, **kw):
# рабочая кнопка для конвертации
    size = (50, 3)
    auto_size_button = True
    path = easygui.fileopenbox(default="HOME/Downloads/*.xls", filetypes='*.xls')
    fname = path
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    sg.popup_ok('начинается проверка условий')
    try:
        wb = excel.Workbooks.Open(fname)
        sg.popup_ok('сконвертировано')
        workbook = openpyxl.load_workbook(wb)
        worksheet = workbook.active
        filetypes = ['*.xlsx', "Excel"]
        default = '*'
        worksheet.unmerge_cells('A1:P1')
        worksheet.unmerge_cells('A2:P2')
        worksheet.unmerge_cells('A3:P3')
        worksheet.unmerge_cells('A4:P4')
        worksheet.unmerge_cells('A5:P5')
        worksheet['D4'] = '= SUBTOTAL(9,D10:D125)'
        worksheet['E4'] = '= SUBTOTAL(9,E10:E125)'
        worksheet['G4'] = '= SUBTOTAL(9,G10:G125)'
        worksheet['F4'].number_format = FORMAT_PERCENTAGE_00
        worksheet['H4'].number_format = FORMAT_PERCENTAGE_00
        worksheet['F4'] = '= E4/D4'
        worksheet['H4'] = '= G4/D4'
        worksheet['A4'] = 'ИТОГИ'
        worksheet['G1'] = 'RNISka Reports 1.2.4'
        worksheet['G1'].font = Font(name='Tahoma', size=9, color="FF0000", italic=True)
        worksheet.merge_cells('D4:D5')
        worksheet.merge_cells('E4:E5')
        worksheet.merge_cells('G4:G5')
        worksheet.merge_cells('F4:F5')
        worksheet.merge_cells('H4:H5')
        worksheet['D4'].alignment = Alignment(horizontal="center", vertical="center")
        worksheet['E4'].alignment = Alignment(horizontal="center", vertical="center")
        worksheet['G4'].alignment = Alignment(horizontal="center", vertical="center")
        worksheet['F4'].alignment = Alignment(horizontal="center", vertical="center")
        worksheet['H4'].alignment = Alignment(horizontal="center", vertical="center")
        worksheet.merge_cells('A4:C5')
        worksheet['A4'].alignment = Alignment(horizontal="left", vertical="center")
        worksheet['D4'].font = Font(bold=True, size=12)
        worksheet['E4'].font = Font(bold=True, size=12)
        worksheet['G4'].font = Font(bold=True, size=12)
        worksheet['F4'].font = Font(bold=True, size=12)
        worksheet['H4'].font = Font(bold=True, size=12)

        worksheet.delete_cols(9, 17)
        ws1 = worksheet.title = "сходы"
        FullRange = "C6:" + get_column_letter(worksheet.max_column) \
                    + str(worksheet.max_row)
        worksheet.auto_filter.ref = FullRange
        ws1 = worksheet.auto_filter.add_filter_column(0, ["173", "177", "179*", "193", "182", "186", "1122", "1123", \
                                                          "2104", "1124", "2811", "1769", "1125", "1131", "1132", \
                                                          "1771", "1127", "1128", "1129", "1130", "1153", "1156", \
                                                          "1134", "1770", "1136", "1137", "1138", "1139", "1141", \
                                                          "1723", "1689", "2185", "1126", "1161", "1172", "1988", \
                                                          "1519", "1697", "3234", "1517", "1518", "1543", "1522", \
                                                          "1524", "1525", "1546", "1526", "1528", "1532"])
        wb.SaveAs(fname + "x", FileFormat=51)  # FileFormat = 51 is for .xlsx extension
        workbook.save(path)
        sg.popup_ok('Файл готов!')
        sg.popup_ok(workbook.sheetnames)
    except Exception as e:
        try:
            wb = excel.Workbooks.Open(fname)

        except Exception as e:
            sg.popup_ok(e)
            sg.popup_ok('неудача')
            wb = None
            return (wb)
    finally:
        sg.popup_ok('завершение')  # задача при которой условия частично выполнены.


def button8():
    sg.popup_ok("принцесса в другом замке епта!")


# кнопочки для проверки белого списка
dispatch_dictionary = {' Егорьевск ':button1, ' Раменское ':button2, ' Шатура ':button3, ' МАП4 ':button4, \
                       ' Конверт ':button5}#, 'сходы':button6, 'ТЕСТ конв':button7, 'ТЕСТ':button8}

# кнопки для конкретно гуи
layout = [[sg.Text('Добро пожаловать!',  auto_size_text=True, justification='center', size=(34,2) )],
          [sg.Text('1.для начала cконвертируйте отчет РНИСа и сохраните его.', auto_size_text=True, justification='center', size=(34,0) )],
          [sg.Text('2. затем выбирайте какие листы создавать. Загрузите файл и затем сохраните.', auto_size_text=True, justification='center', size=(34,3) )],
          [sg.Button(' Егорьевск ', size=(10,2), border_width=(5), pad=((0, 0), 0)), sg.Button(' Раменское ', size=(10,2), border_width=(5), pad=((0, 0), 0)), \
           sg.Button(' Шатура ', size=(10,2), border_width=(5), pad=((0, 0), 0))],
          [sg.Button(' МАП4 ', size=(34,2), border_width=(5), pad=((0, 0), 0)) ],
          [sg.Button(' Конверт ', size=(34,2), border_width=(5), pad=((0, 0), 0))],
#          [sg.Button('сходы', size=(6,2), pad=((0, 0), 0))],
#          [sg.Button('ТЕСТ конв', size=(10,2), border_width=(5), pad=((0, 0), 0))],
#          [sg.Button('ТЕСТ', size=(10,2), border_width=(5), pad=((0, 0), 0))],
#          [sg.Text('_' * 100, size=(65, 1))],
#              [sg.Text('Flags', font=('Helvetica', 15), justification='left')],
#              [sg.Checkbox('Егорьевск', size=(12, 1), default=True), sg.Checkbox('Раменское', size=(20, 1))],
#              [sg.Checkbox('Шатура', size=(12, 1)), sg.Checkbox('Только вывести на экран', size=(20, 1), default=True)],
#              [sg.Checkbox('МАП №4', size=(12, 1)), sg.Checkbox('Keep Intermediate Data', size=(20, 1))],
#              [sg.Text('_' * 100, size=(65, 1))],
#          [sg.FileBrowse(file_types=(("Excel", "*.xls"),))],
          [sg.Quit(' Выход ', size=(6,2), pad=((230, 0), 3))]]

# титульное окно
window = sg.Window('RNISka Reports 1.2.4', layout)

# белый список
while True:
    # ифелс
    event, value = window.read()
    if event in (' Выход ', sg.WIN_CLOSED):
        break
    # белый список в действии
    if event in dispatch_dictionary:
        func_to_call = dispatch_dictionary[event]   # словарь в действии
        func_to_call()
    else:
        sg.Print('Возвращаемое событие {} отсутствует в словаре'.format(event))

window.close()
#    все.
