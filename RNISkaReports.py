import PySimpleGUI as sg
import easygui
import openpyxl
from openpyxl.utils import get_column_letter
import win32com.client as win32


sg.theme('DarkBlue9')
# This design pattern simulates button callbacks
# This implementation uses a simple "Dispatch Dictionary" to store events and functions

# The callback functions


def button1():
    path = easygui.fileopenbox(default= '*.xlsx', filetypes= '*.xlsx')
    path2 = easygui.filesavebox(default= 'Egoryevsk.xlsx', filetypes= '*.xslx')
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook["Лист 1"]
    filetypes = ['*.xlsx', "Excel"]
    default = '*'
    worksheet = workbook["Лист 1"]
    worksheet.delete_cols (9,17)
    ws1 = worksheet.title = "Егорьевск"
    FullRange = "C6:" + get_column_letter(worksheet.max_column) \
    + str(worksheet.max_row)
    worksheet.auto_filter.ref = FullRange
    ws1 = worksheet.auto_filter.add_filter_column(0, ["167", "168", "169", "171", "172", "173", "174", "175", "176", \
                                                      "177", "178", "180", "181", "182", "183", "184", "185", "186",\
                                                      "187", "188", "189", "190", "193", "194", "678", "1853", "1855",\
                                                      "2102", "2600", "3252", "3253", "3254", "179*"])
    workbook.save(path2)
    sg.popup_ok('Файл готов!')

def button2():
    path = easygui.fileopenbox(default= '*.xlsx', filetypes= '*.xlsx')
    path2 = easygui.filesavebox(default= 'Ramenskoye.xlsx', filetypes= '*.xslx')
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook["Лист 1"]
    filetypes=['*.xlsx', "Excel"]
    default='*'
    worksheet = workbook["Лист 1"]
    worksheet.delete_cols (9,17)
    ws1 = worksheet.title = "Раменское"
    FullRange = "C6:" + get_column_letter(worksheet.max_column) \
    + str(worksheet.max_row)
    worksheet.auto_filter.ref = FullRange
    ws1 = worksheet.auto_filter.add_filter_column(0, ["1122", "1123", "1124", "1125", "1126", "1127", "1128", "1129", "1130", "1131", "1132", "1133", "1134", "1135", "1136", "1137", "1138", "1139", "1140", "1141", "1142", "1147", "1153", "1154", "1156", "1161", "1689", "1723", "1724", "1769", "1770", "1771", "2104", "2120", "2185", "2811", "3063", "3066", "3067", "3074"])
    workbook.save(path2)
    sg.popup_ok('Файл готов!')

def button3():
    path = easygui.fileopenbox(default= '*.xlsx', filetypes= '*.xlsx')
    path2 = easygui.filesavebox(default= 'Shatoora.xlsx', filetypes= '*.xslx')
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook["Лист 1"]
    filetypes=['*.xlsx', "Excel"]
    default='*'
    worksheet = workbook["Лист 1"]
    worksheet.delete_cols (9,17)
    ws1 = worksheet.title = "Шатура"
    FullRange = "C6:" + get_column_letter(worksheet.max_column) \
    + str(worksheet.max_row)
    worksheet.auto_filter.ref = FullRange
    ws1 = worksheet.auto_filter.add_filter_column(0, ["1513", "1172", "1993", "1514", "1515", "1516", "1988", "1519", "1697", "3234", "2082", "1517", "1518", "1521", "1543", "1522", "2491", "1523", "1524", "1525", "1544", "2609", "1546", "1547", "1526", "1527", "1528", "1529", "1530", "1548", "1531", "1532", "1550", "2565", "1549", "1545", "1991", "1896", "1534", "1538", "1539", "1811", "1540", "1541", "1812", "1542", "1759"])
    workbook.save(path2)
    sg.popup_ok('Файл готов!')

def button4():
    path = easygui.fileopenbox(default= '*.xlsx', filetypes= '*.xlsx')
    path2 = easygui.filesavebox(default= 'map4.xlsx', filetypes= '*.xslx')
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook["Лист 1"]
    filetypes=['*.xlsx', "Excel"]
    default='*'
    worksheet = workbook["Лист 1"]
    worksheet.delete_cols (9,17)
    ws1 = worksheet.title = "Егорьевск"
    FullRange = "C6:" + get_column_letter(worksheet.max_column) \
    + str(worksheet.max_row)
    worksheet.auto_filter.ref = FullRange
    ws1 = worksheet.auto_filter.add_filter_column(0, ["167", "168", "169", "171", "172", "173", "174", "175", "176", "177", "178", "180", "181", "182", "183", "184", "185", "186", "187", "188", "189", "190", "193", "194", "678", "1853", "1855", "2102", "2600", "3252", "3253", "3254", "179*"])

    source = workbook.active
    target = workbook.copy_worksheet(source)
    worksheet = workbook["Егорьевск Copy"]
    ws2 = worksheet.title = "Раменское"
    FullRange = "C6:" + get_column_letter(worksheet.max_column) \
    + str(worksheet.max_row)
    worksheet.auto_filter.ref = FullRange
    ws2 = worksheet.auto_filter.add_filter_column(0, ["1122", "1123", "1124", "1125", "1126", "1127", "1128", "1129", "1130", "1131", "1132", "1133", "1134", "1135", "1136", "1137", "1138", "1139", "1140", "1141", "1142", "1147", "1153", "1154", "1156", "1161", "1689", "1723", "1724", "1769", "1770", "1771", "2104", "2120", "2185", "2811", "3063", "3066", "3067", "3074"])

    source = workbook.active
    target = workbook.copy_worksheet(source)
    worksheet = workbook["Егорьевск Copy"]
    ws3 = worksheet.title = "Шатура"
    FullRange = "C6:" + get_column_letter(worksheet.max_column) \
    + str(worksheet.max_row)
    worksheet.auto_filter.ref = FullRange
    ws3 = worksheet.auto_filter.add_filter_column(0, ["1513", "1172", "1993", "1514", "1515", "1516", "1988", "1519", "1697", "3234", "2082", "1517", "1518", "1521", "1543", "1522", "2491", "1523", "1524", "1525", "1544", "2609", "1546", "1547", "1526", "1527", "1528", "1529", "1530", "1548", "1531", "1532", "1550", "2565", "1549", "1545", "1991", "1896", "1534", "1538", "1539", "1811", "1540", "1541", "1812", "1542", "1759"])
    workbook.save(path2)
    sg.popup_ok('Файл готов!')

def button5():
        size=(50,3)
        auto_size_button=True
        path = easygui.fileopenbox(default= '*.xls', filetypes= '*.xls')
        fname = path
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(fname)
        wb.SaveAs(fname+"x", FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
        wb.Close()                               #FileFormat = 56 is for .xls extension
        excel.Application.Quit()
        sg.popup_ok ('Успешно сконвертировано!')


# кнопочки для проверки белого списка
dispatch_dictionary = {' Егорьевск ':button1, ' Раменское ':button2, ' Шатура ':button3, ' МАП4 ':button4, ' Конверт ':button5}

# кнопки для конкретно гуи
layout = [[sg.Text('Добро пожаловать!',  auto_size_text=True, justification='center', size=(34,2) )],
          [sg.Text('1.для начала cконвертируйте отчет РНИСа и сохраните его.', auto_size_text=True, justification='center', size=(34,0) )],
          [sg.Text('2. затем выбирайте какие листы создавать. Загрузите файл и затем сохраните.', auto_size_text=True, justification='center', size=(34,3) )],
          [sg.Button(' Егорьевск ', size=(10,2), border_width=(5), pad=((0, 0), 0)), sg.Button(' Раменское ', size=(10,2), border_width=(5), pad=((0, 0), 0)), \
           sg.Button(' Шатура ', size=(10,2), border_width=(5), pad=((0, 0), 0))],
          [sg.Button(' МАП4 ', size=(34,2), border_width=(5), pad=((0, 0), 0)) ], \
          [sg.Button(' Конверт ', size=(34,2), border_width=(5), pad=((0, 0), 0))],
          [sg.Quit(' Выход ', size=(6,2), pad=((230, 0), 3))]]

# титульное окно
window = sg.Window('RNISka Reports', layout)

# белый список
while True:
    # ифелс
    event, value = window.read()
    if event in (' Выход ', sg.WIN_CLOSED):
        break
    # белый список в действии
    if event in dispatch_dictionary:
        func_to_call = dispatch_dictionary[event]   # словарь в действии
        func_to_call()
    else:
        sg.Print('Возвращаемое событие {} отсутствует в словаре'.format(event))



window.close()

    # все.
