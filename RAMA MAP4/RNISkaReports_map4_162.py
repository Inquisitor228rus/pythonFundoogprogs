import os
import re
import shutil
import tempfile
import time
from datetime import date

import PySimpleGUI as sg
import easygui
import imwatchingyou
import openpyxl
import win32com.client as win32
from openpyxl.styles import Alignment, Font
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter

today = date.today()
sg.theme('DarkBlue9')

file_types = [("Excel (*.xlsx)", "*.xlsx"), ("All files (*.*)", "*.*")]
tmp_file = tempfile.NamedTemporaryFile(suffix=".xlsx").name

versionRR = " 1.6.2"
nameProg = "RNISka Reports"
betaOrNot = ""
mapEdit = "МАП-4"
testWindows = nameProg + " " + versionRR + " " + mapEdit + " Edition"

WIN_W: int = 12
WIN_H: int = 3
Pause_Sleep: float = 0

##########           Егыч СУММА
ego1 = '= SUBTOTAL(9,D8:D119)'
ego2 = '= SUBTOTAL(9,E8:E119)'
ego3 = '= SUBTOTAL(9,G8:G119)'
##########           Раменское СУММА
rama1 = '= SUBTOTAL(9,D10:D125)'
rama2 = '= SUBTOTAL(9,E10:E125)'
rama3 = '= SUBTOTAL(9,G10:G125)'
##########           Шатурка СУММА
shatoor1 = '= SUBTOTAL(9,D9:D127)'  # Коломна город
shatoor2 = '= SUBTOTAL(9,E9:E127)'
shatoor3 = '= SUBTOTAL(9,G9:G127)'

egor = ["167", "168", "169", "171", "172", "173", "174", "175", "176", \
                                                      "177", "178", "180", "181", "182", "183", "184", "185", "186", \
                                                      "187", "188", "189", "190", "193", "194", "678", "1853", "1855", \
                                                      "2102", "2600", "3252", "3253", "3254", "179*"]
ramen = ["1122", "1123", "1124", "1125", "1126", "1127", "1128", "1129", \
                                                      "1130", "1131", "1132", "1133", "1134", "1135", "1136", "1137", \
                                                      "1138", "1139", "1140", "1141", "1142", "1147", "1153", "1154", \
                                                      "1156", "1161", "1689", "1723", "1724", "1769", "1770", "1771", \
                                                      "2104", "2120", "2185", "2811", "3063", "3066", "3067", "3074"]
shatoora = ["1513", "1172", "1993", "1514", "1515", "1516", "1988", "1519", \
                                                      "1697", "3234", "2082", "1517", "1518", "1521", "1543", "1522", \
                                                      "2491", "1523", "1524", "1525", "1544", "2609", "1546", "1547", \
                                                      "1526", "1527", "1528", "1529", "1530", "1548", "1531", "1532", \
                                                      "1550", "2565", "1549", "1545", "1991", "1896", "1534", "1538", \
                                                      "1539", "1811", "1540", "1541", "1812", "1542", "1759"]

def button1():
    progress_bar.UpdateBar(0, 20)
    time.sleep(Pause_Sleep)
    path = easygui.fileopenbox(default="HOME/Downloads/*.xlsx", filetypes='*.xlsx')
    progress_bar.UpdateBar(1, 20)
    time.sleep(Pause_Sleep)
    try:
        if path and os.path.exists(path):
            shutil.copy(path, tmp_file)
            workbook = openpyxl.load_workbook(tmp_file)
    except Exception as e:
        try:
            if path and os.path.exists(path):
                shutil.copy(path, tmp_file)
                workbook = openpyxl.load_workbook(tmp_file)
        except Exception as e:
            sg.popup_auto_close(e, 'Ошибка при открытии...')
            #sg.popup_ok(e)
            workbook = None
            progress_bar.UpdateBar(0, 20)
        return (workbook)
    progress_bar.UpdateBar(2, 20)
    time.sleep(Pause_Sleep)

    progress_bar.UpdateBar(3, 20)
    time.sleep(Pause_Sleep)
    try:
        worksheet = workbook["Лист 1"]
    except Exception as e:
        try:
            worksheet = workbook["Лист 1"]
        except Exception as e:
            sg.popup_auto_close(e,
                'Возможно выбран не тот файл. \nВыберете "Итоговый отчет о работе выходов по филиалу (общие показатели)"')
            #sg.popup_auto_close(e)
            worksheet = None
            progress_bar.UpdateBar(0, 20)
        return (worksheet)
    filetypes = ['*.xlsx', "Excel"]
    default = '*'
    progress_bar.UpdateBar(4, 20)
    time.sleep(Pause_Sleep)
    dataSheet = worksheet['A2']
    tosafe = re.findall(r'\d{2}.\d{2}.\d{2}', dataSheet.value)

    worksheet.unmerge_cells('A1:P1')
    worksheet.unmerge_cells('A2:P2')
    worksheet.unmerge_cells('A3:P3')
    worksheet.unmerge_cells('A4:P4')
    worksheet.unmerge_cells('A5:P5')
    progress_bar.UpdateBar(5, 20)
    time.sleep(Pause_Sleep)
    worksheet['D4'] = ego1
    worksheet['E4'] = ego2
    worksheet['G4'] = ego3
    progress_bar.UpdateBar(6, 20)
    time.sleep(Pause_Sleep)
    worksheet['F4'].number_format = FORMAT_PERCENTAGE_00
    worksheet['H4'].number_format = FORMAT_PERCENTAGE_00
    progress_bar.UpdateBar(7, 20)
    time.sleep(Pause_Sleep)
    worksheet['F4'] = '= E4/D4'
    worksheet['H4'] = '= G4/D4'
    progress_bar.UpdateBar(8, 20)
    time.sleep(Pause_Sleep)
    worksheet['A4'] = 'ИТОГИ'
    worksheet['C4'] = 'Егорьевск'
    worksheet['G1'] = nameProg + versionRR + betaOrNot
    progress_bar.UpdateBar(9, 20)
    time.sleep(Pause_Sleep)
    worksheet['G1'].font = Font(name='Tahoma', size=9, color="FF0000", italic=True)
    progress_bar.UpdateBar(10, 20)
    time.sleep(Pause_Sleep)
    worksheet.merge_cells('D4:D5')
    worksheet.merge_cells('E4:E5')
    worksheet.merge_cells('G4:G5')
    worksheet.merge_cells('F4:F5')
    worksheet.merge_cells('H4:H5')
    progress_bar.UpdateBar(11, 20)
    time.sleep(Pause_Sleep)
    worksheet['D4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['E4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['G4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['F4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['H4'].alignment = Alignment(horizontal="center", vertical="center")
    progress_bar.UpdateBar(12, 20)
    time.sleep(Pause_Sleep)
    worksheet.merge_cells('A4:B5')
    worksheet.merge_cells('C4:C5')
    progress_bar.UpdateBar(13, 20)
    time.sleep(Pause_Sleep)
    worksheet['A4'].alignment = Alignment(horizontal="left", vertical="center")
    worksheet['C4'].font = Font(bold=True, size=12)
    worksheet['C4'].alignment = Alignment(horizontal="center", vertical="center")
    progress_bar.UpdateBar(14, 20)
    time.sleep(Pause_Sleep)
    worksheet['D4'].font = Font(bold=True, size=12)
    worksheet['E4'].font = Font(bold=True, size=12)
    worksheet['G4'].font = Font(bold=True, size=12)
    worksheet['F4'].font = Font(bold=True, size=12)
    worksheet['H4'].font = Font(bold=True, size=12)
    progress_bar.UpdateBar(15, 20)
    time.sleep(Pause_Sleep)
    worksheet.delete_cols(9, 17)
    progress_bar.UpdateBar(16, 20)
    time.sleep(Pause_Sleep)
    ws1 = worksheet.title = "Егорьевск"
    progress_bar.UpdateBar(17, 20)
    time.sleep(Pause_Sleep)
    FullRange = "C6:" + get_column_letter(worksheet.max_column) \
                + str(worksheet.max_row)
    progress_bar.UpdateBar(18, 20)
    time.sleep(Pause_Sleep)
    worksheet.auto_filter.ref = FullRange
    ws1 = worksheet.auto_filter.add_filter_column(0, egor)
    progress_bar.UpdateBar(19, 20)
    time.sleep(Pause_Sleep)
    safepath = today.strftime("%d,%m,%Y" + ".xlsx")
    save_filename = value["folder_s"] + "/Egoryevsk_" + tosafe[0] + "_" + tosafe[1] + ".xlsx"# + re.findall(r'\d{2}.\d{2}.\d{2}', dataSheet.value)

    if os.path.isfile(save_filename):
        expand = 1
        while True:
            expand += 1
            new_file_name = save_filename.split(".xlsx")[0] + " (" +str(expand) + ")" + ".xlsx"
            if os.path.isfile(new_file_name):
                continue
            else:
                save_filename = new_file_name
                break
    try:
        #save_filename = path

        #path2 =
        #sg.popup(f"Saved: {save_filename}")
        #path2 = easygui.filesavebox(default="HOME/APPDATA/RPBeta/Egoryevsk.xlsx", filetypes='*.xslx')
        workbook.save(tmp_file)
        shutil.copy(tmp_file, save_filename)
    except Exception as e:
        try:
            workbook.save(tmp_file)
            shutil.copy(tmp_file, save_filename)
        except Exception as e:
            sg.popup_auto_close(e, 'Вы отказались от сохранения. \nНо файл сохранился в той же папке.')
            #sg.popup_ok(e)
            workbook.save(path + safepath)
            progress_bar.UpdateBar(0, 20)
        return (safepath)
    progress_bar.UpdateBar(20, 20)
    time.sleep(Pause_Sleep)
    sg.popup_ok('Файл готов!')
    progress_bar.UpdateBar(0, 20)


def button2():
    progress_bar.UpdateBar(0, 20)
    time.sleep(Pause_Sleep)
    path = easygui.fileopenbox(default="HOME/Downloads/*.xlsx", filetypes='*.xlsx')
    progress_bar.UpdateBar(1, 20)
    time.sleep(Pause_Sleep)
    try:
        if path and os.path.exists(path):
            shutil.copy(path, tmp_file)
            workbook = openpyxl.load_workbook(tmp_file)
    except Exception as e:
        try:
            if path and os.path.exists(path):
                shutil.copy(path, tmp_file)
                workbook = openpyxl.load_workbook(tmp_file)
        except Exception as e:
            sg.popup_auto_close(e, 'Ошибка при открытии...')

            workbook = None
            progress_bar.UpdateBar(0, 20)
        return (workbook)
    progress_bar.UpdateBar(2, 20)
    time.sleep(Pause_Sleep)

    progress_bar.UpdateBar(3, 20)
    time.sleep(Pause_Sleep)
    try:
        worksheet = workbook["Лист 1"]
    except Exception as e:
        try:
            worksheet = workbook["Лист 1"]
        except Exception as e:
            sg.popup_auto_close(
                'Возможно выбран не тот файл. \nВыберете "Итоговый отчет о работе выходов по филиалу (общие показатели)"')
            sg.popup_auto_close(e)
            worksheet = None
            progress_bar.UpdateBar(0, 20)
        return (worksheet)
    filetypes = ['*.xlsx', "Excel"]
    default = '*'
    progress_bar.UpdateBar(4, 20)
    time.sleep(Pause_Sleep)
    dataSheet = worksheet['A2']
    tosafe = re.findall(r'\d{2}.\d{2}.\d{2}', dataSheet.value)

    worksheet.unmerge_cells('A1:P1')
    worksheet.unmerge_cells('A2:P2')
    worksheet.unmerge_cells('A3:P3')
    worksheet.unmerge_cells('A4:P4')
    worksheet.unmerge_cells('A5:P5')
    progress_bar.UpdateBar(5, 20)
    time.sleep(Pause_Sleep)
    worksheet['D4'] = rama1
    worksheet['E4'] = rama2
    worksheet['G4'] = rama3
    progress_bar.UpdateBar(6, 20)
    time.sleep(Pause_Sleep)
    worksheet['F4'].number_format = FORMAT_PERCENTAGE_00
    worksheet['H4'].number_format = FORMAT_PERCENTAGE_00
    progress_bar.UpdateBar(7, 20)
    time.sleep(Pause_Sleep)
    worksheet['F4'] = '= E4/D4'
    worksheet['H4'] = '= G4/D4'
    progress_bar.UpdateBar(8, 20)
    time.sleep(Pause_Sleep)
    worksheet['A4'] = 'ИТОГИ'
    worksheet['C4'] = 'Раменское'
    worksheet['G1'] = nameProg + versionRR + betaOrNot
    progress_bar.UpdateBar(9, 20)
    time.sleep(Pause_Sleep)
    worksheet['G1'].font = Font(name='Tahoma', size=9, color="FF0000", italic=True)
    progress_bar.UpdateBar(10, 20)
    time.sleep(Pause_Sleep)
    worksheet.merge_cells('D4:D5')
    worksheet.merge_cells('E4:E5')
    worksheet.merge_cells('G4:G5')
    worksheet.merge_cells('F4:F5')
    worksheet.merge_cells('H4:H5')
    progress_bar.UpdateBar(11, 20)
    time.sleep(Pause_Sleep)
    worksheet['D4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['E4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['G4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['F4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['H4'].alignment = Alignment(horizontal="center", vertical="center")
    progress_bar.UpdateBar(12, 20)
    time.sleep(Pause_Sleep)
    worksheet.merge_cells('A4:B5')
    worksheet.merge_cells('C4:C5')
    progress_bar.UpdateBar(13, 20)
    time.sleep(Pause_Sleep)
    worksheet['A4'].alignment = Alignment(horizontal="left", vertical="center")
    worksheet['C4'].font = Font(bold=True, size=12)
    worksheet['C4'].alignment = Alignment(horizontal="center", vertical="center")
    progress_bar.UpdateBar(14, 20)
    time.sleep(Pause_Sleep)
    worksheet['D4'].font = Font(bold=True, size=12)
    worksheet['E4'].font = Font(bold=True, size=12)
    worksheet['G4'].font = Font(bold=True, size=12)
    worksheet['F4'].font = Font(bold=True, size=12)
    worksheet['H4'].font = Font(bold=True, size=12)
    progress_bar.UpdateBar(15, 20)
    time.sleep(Pause_Sleep)
    worksheet.delete_cols(9, 17)
    progress_bar.UpdateBar(16, 20)
    time.sleep(Pause_Sleep)
    ws1 = worksheet.title = "Раменское"
    progress_bar.UpdateBar(17, 20)
    time.sleep(Pause_Sleep)
    FullRange = "C6:" + get_column_letter(worksheet.max_column) \
                + str(worksheet.max_row)
    progress_bar.UpdateBar(18, 20)
    time.sleep(Pause_Sleep)
    worksheet.auto_filter.ref = FullRange
    ws1 = worksheet.auto_filter.add_filter_column(0, ramen)
    progress_bar.UpdateBar(19, 20)
    time.sleep(Pause_Sleep)
    safepath = today.strftime("%d,%m,%Y" + ".xlsx")
    save_filename = value["folder_s"] + "/Ramenskoye_" + tosafe[0] + "_" + tosafe[
        1] + ".xlsx"  # + re.findall(r'\d{2}.\d{2}.\d{2}', dataSheet.value)

    if os.path.isfile(save_filename):
        expand = 1
        while True:
            expand += 1
            new_file_name = save_filename.split(".xlsx")[0] + " (" + str(expand) + ")" + ".xlsx"
            if os.path.isfile(new_file_name):
                continue
            else:
                save_filename = new_file_name
                break
    try:
        # save_filename = path

        # path2 =
        # sg.popup(f"Saved: {save_filename}")
        # path2 = easygui.filesavebox(default="HOME/APPDATA/RPBeta/Egoryevsk.xlsx", filetypes='*.xslx')
        workbook.save(tmp_file)
        shutil.copy(tmp_file, save_filename)
    except Exception as e:
        try:
            workbook.save(tmp_file)
            shutil.copy(tmp_file, save_filename)
        except Exception as e:
            sg.popup_auto_close(e, 'Вы отказались от сохранения. \nНо файл сохранился в той же папке.')
            # sg.popup_ok(e)
            workbook.save(path + safepath)
            progress_bar.UpdateBar(0, 20)
        return (safepath)
    progress_bar.UpdateBar(20, 20)
    time.sleep(Pause_Sleep)
    sg.popup_ok('Файл готов!')
    progress_bar.UpdateBar(0, 20)


def button3():
    progress_bar.UpdateBar(0, 20)
    time.sleep(Pause_Sleep)
    path = easygui.fileopenbox(default="HOME/Downloads/*.xlsx", filetypes='*.xlsx')
    progress_bar.UpdateBar(1, 20)
    time.sleep(Pause_Sleep)
    try:
        if path and os.path.exists(path):
            shutil.copy(path, tmp_file)
            workbook = openpyxl.load_workbook(tmp_file)
    except Exception as e:
        try:
            if path and os.path.exists(path):
                shutil.copy(path, tmp_file)
                workbook = openpyxl.load_workbook(tmp_file)
        except Exception as e:
            sg.popup_auto_close(e, 'Ошибка при открытии...')
            # sg.popup_ok(e)
            workbook = None
            progress_bar.UpdateBar(0, 20)
        return (workbook)
    progress_bar.UpdateBar(2, 20)
    time.sleep(Pause_Sleep)

    progress_bar.UpdateBar(3, 20)
    time.sleep(Pause_Sleep)
    try:
        worksheet = workbook["Лист 1"]
    except Exception as e:
        try:
            worksheet = workbook["Лист 1"]
        except Exception as e:
            sg.popup_auto_close(e,
                'Возможно выбран не тот файл. \nВыберете "Итоговый отчет о работе выходов по филиалу (общие показатели)"')
            #sg.popup_auto_close(e)
            worksheet = None
            progress_bar.UpdateBar(0, 20)
        return (worksheet)
    filetypes = ['*.xlsx', "Excel"]
    default = '*'
    progress_bar.UpdateBar(4, 20)
    time.sleep(Pause_Sleep)
    dataSheet = worksheet['A2']
    tosafe = re.findall(r'\d{2}.\d{2}.\d{2}', dataSheet.value)
    worksheet.unmerge_cells('A1:P1')
    worksheet.unmerge_cells('A2:P2')
    worksheet.unmerge_cells('A3:P3')
    worksheet.unmerge_cells('A4:P4')
    worksheet.unmerge_cells('A5:P5')
    progress_bar.UpdateBar(5, 20)
    time.sleep(Pause_Sleep)
    worksheet['D4'] = shatoor1
    worksheet['E4'] = shatoor2
    worksheet['G4'] = shatoor3
    progress_bar.UpdateBar(6, 20)
    time.sleep(Pause_Sleep)
    worksheet['F4'].number_format = FORMAT_PERCENTAGE_00
    worksheet['H4'].number_format = FORMAT_PERCENTAGE_00
    progress_bar.UpdateBar(7, 20)
    time.sleep(Pause_Sleep)
    worksheet['F4'] = '= E4/D4'
    worksheet['H4'] = '= G4/D4'
    progress_bar.UpdateBar(8, 20)
    time.sleep(Pause_Sleep)
    worksheet['A4'] = 'ИТОГИ'
    worksheet['C4'] = 'Шатура'
    worksheet['G1'] = nameProg + versionRR + betaOrNot
    progress_bar.UpdateBar(9, 20)
    time.sleep(Pause_Sleep)
    worksheet['G1'].font = Font(name='Tahoma', size=9, color="FF0000", italic=True)
    progress_bar.UpdateBar(10, 20)
    time.sleep(Pause_Sleep)
    worksheet.merge_cells('D4:D5')
    worksheet.merge_cells('E4:E5')
    worksheet.merge_cells('G4:G5')
    worksheet.merge_cells('F4:F5')
    worksheet.merge_cells('H4:H5')
    progress_bar.UpdateBar(11, 20)
    time.sleep(Pause_Sleep)
    worksheet['D4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['E4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['G4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['F4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['H4'].alignment = Alignment(horizontal="center", vertical="center")
    progress_bar.UpdateBar(12, 20)
    time.sleep(Pause_Sleep)
    worksheet.merge_cells('A4:B5')
    worksheet.merge_cells('C4:C5')
    progress_bar.UpdateBar(13, 20)
    time.sleep(Pause_Sleep)
    worksheet['A4'].alignment = Alignment(horizontal="left", vertical="center")
    worksheet['C4'].font = Font(bold=True, size=12)
    worksheet['C4'].alignment = Alignment(horizontal="center", vertical="center")
    progress_bar.UpdateBar(14, 20)
    time.sleep(Pause_Sleep)
    worksheet['D4'].font = Font(bold=True, size=12)
    worksheet['E4'].font = Font(bold=True, size=12)
    worksheet['G4'].font = Font(bold=True, size=12)
    worksheet['F4'].font = Font(bold=True, size=12)
    worksheet['H4'].font = Font(bold=True, size=12)
    progress_bar.UpdateBar(15, 20)
    time.sleep(Pause_Sleep)
    worksheet.delete_cols(9, 17)
    progress_bar.UpdateBar(16, 20)
    time.sleep(Pause_Sleep)
    ws1 = worksheet.title = "Шатура"
    progress_bar.UpdateBar(17, 20)
    time.sleep(Pause_Sleep)
    FullRange = "C6:" + get_column_letter(worksheet.max_column) \
                + str(worksheet.max_row)
    progress_bar.UpdateBar(18, 20)
    time.sleep(Pause_Sleep)
    worksheet.auto_filter.ref = FullRange
    ws1 = worksheet.auto_filter.add_filter_column(0, shatoora)
    progress_bar.UpdateBar(19, 20)
    time.sleep(Pause_Sleep)
    safepath = today.strftime("%d,%m,%Y" + ".xlsx")
    save_filename = value["folder_s"] + "/Shatoora_" + tosafe[0] + "_" + tosafe[
        1] + ".xlsx"  # + re.findall(r'\d{2}.\d{2}.\d{2}', dataSheet.value)

    if os.path.isfile(save_filename):
        expand = 1
        while True:
            expand += 1
            new_file_name = save_filename.split(".xlsx")[0] + " (" + str(expand) + ")" + ".xlsx"
            if os.path.isfile(new_file_name):
                continue
            else:
                save_filename = new_file_name
                break
    try:
        # save_filename = path

        # path2 =
        # sg.popup(f"Saved: {save_filename}")
        # path2 = easygui.filesavebox(default="HOME/APPDATA/RPBeta/Egoryevsk.xlsx", filetypes='*.xslx')
        workbook.save(tmp_file)
        shutil.copy(tmp_file, save_filename)
    except Exception as e:
        try:
            workbook.save(tmp_file)
            shutil.copy(tmp_file, save_filename)
        except Exception as e:
            sg.popup_auto_close(e, 'Вы отказались от сохранения. \nНо файл сохранился в той же папке.')
            # sg.popup_ok(e)
            workbook.save(path + safepath)
            progress_bar.UpdateBar(0, 20)
        return (safepath)
    progress_bar.UpdateBar(20, 20)
    time.sleep(Pause_Sleep)
    sg.popup_ok('Файл готов!')
    progress_bar.UpdateBar(0, 20)


def button4():
    progress_bar.UpdateBar(0, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    path = easygui.fileopenbox(default="HOME/Downloads/*.xlsx", filetypes='*.xlsx')
    try:
        if path and os.path.exists(path):
            shutil.copy(path, tmp_file)
            workbook = openpyxl.load_workbook(tmp_file)
    except Exception as e:
        try:
            if path and os.path.exists(path):
                shutil.copy(path, tmp_file)
                workbook = openpyxl.load_workbook(tmp_file)
        except Exception as e:
            sg.popup_auto_close(e, 'Ошибка при открытии...')
            # sg.popup_ok(e)
            workbook = None
            progress_bar.UpdateBar(0, 33)
        return (workbook)
    progress_bar.UpdateBar(1, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)

    progress_bar.UpdateBar(2, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    try:
        worksheet = workbook["Лист 1"]
    except Exception as e:
        try:
            worksheet = workbook["Лист 1"]
        except Exception as e:
            sg.popup_auto_close(e,
                'Возможно выбран не тот файл. \nВыберете "Итоговый отчет о работе выходов по филиалу (общие показатели)"')
            #sg.popup_auto_close(e)
            worksheet = None
            progress_bar.UpdateBar(0, 33)
        return (worksheet)
    filetypes = ['*.xlsx', "Excel"]
    default = '*'
    worksheet = workbook["Лист 1"]
    progress_bar.UpdateBar(3, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    dataSheet = worksheet['A2']
    tosafe = re.findall(r'\d{2}.\d{2}.\d{2}', dataSheet.value)
    worksheet.unmerge_cells('A1:P1')
    worksheet.unmerge_cells('A2:P2')
    worksheet.unmerge_cells('A3:P3')
    worksheet.unmerge_cells('A4:P4')
    worksheet.unmerge_cells('A5:P5')
    progress_bar.UpdateBar(4, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet['D4'] = ego1
    worksheet['E4'] = ego2
    worksheet['G4'] = ego3
    progress_bar.UpdateBar(5, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet['F4'].number_format = FORMAT_PERCENTAGE_00
    worksheet['H4'].number_format = FORMAT_PERCENTAGE_00
    progress_bar.UpdateBar(6, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet['F4'] = '= E4/D4'
    worksheet['H4'] = '= G4/D4'
    progress_bar.UpdateBar(7, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet['A4'] = 'ИТОГИ'
    worksheet['C4'] = 'Егорьевск'
    worksheet['G1'] = nameProg + versionRR + betaOrNot
    progress_bar.UpdateBar(8, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet['G1'].font = Font(name='Tahoma', size=9, color="FF0000", italic=True)
    progress_bar.UpdateBar(9, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet.merge_cells('D4:D5')
    worksheet.merge_cells('E4:E5')
    worksheet.merge_cells('G4:G5')
    worksheet.merge_cells('F4:F5')
    worksheet.merge_cells('H4:H5')
    progress_bar.UpdateBar(10, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet['D4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['E4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['G4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['F4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['H4'].alignment = Alignment(horizontal="center", vertical="center")
    progress_bar.UpdateBar(11, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet.merge_cells('A4:B5')
    worksheet.merge_cells('C4:C5')
    progress_bar.UpdateBar(12, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet['A4'].alignment = Alignment(horizontal="left", vertical="center")
    worksheet['C4'].font = Font(bold=True, size=12)
    worksheet['C4'].alignment = Alignment(horizontal="center", vertical="center")
    progress_bar.UpdateBar(13, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet['D4'].font = Font(bold=True, size=12)
    worksheet['E4'].font = Font(bold=True, size=12)
    worksheet['G4'].font = Font(bold=True, size=12)
    worksheet['F4'].font = Font(bold=True, size=12)
    worksheet['H4'].font = Font(bold=True, size=12)
    progress_bar.UpdateBar(14, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet.delete_cols(9, 17)
    progress_bar.UpdateBar(15, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    ws1 = worksheet.title = "Егорьевск"
    progress_bar.UpdateBar(16, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    FullRange = "C6:" + get_column_letter(worksheet.max_column) \
                + str(worksheet.max_row)
    progress_bar.UpdateBar(17, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet.auto_filter.ref = FullRange
    progress_bar.UpdateBar(18, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    ws1 = worksheet['D4'] = ego1
    ws1 = worksheet['E4'] = ego2
    ws1 = worksheet['G4'] = ego3
    ws1 = worksheet['C4'] = 'Егорьевск'
    ws1 = worksheet.auto_filter.add_filter_column(0, egor)
    progress_bar.UpdateBar(19, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    source = workbook.active
    progress_bar.UpdateBar(20, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    target = workbook.copy_worksheet(source)
    progress_bar.UpdateBar(21, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet = workbook["Егорьевск Copy"]
    progress_bar.UpdateBar(22, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    ws2 = worksheet.title = "Раменское"
    progress_bar.UpdateBar(23, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    FullRange = "C6:" + get_column_letter(worksheet.max_column) \
                + str(worksheet.max_row)
    progress_bar.UpdateBar(24, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet.auto_filter.ref = FullRange
    progress_bar.UpdateBar(25, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    ws2 = worksheet['D4'] = rama1
    ws2 = worksheet['E4'] = rama2
    ws2 = worksheet['G4'] = rama3
    ws2 = worksheet['C4'] = 'Раменское'
    ws2 = worksheet.auto_filter.add_filter_column(0, ramen)
    progress_bar.UpdateBar(26, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    source = workbook.active
    progress_bar.UpdateBar(27, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    target = workbook.copy_worksheet(source)
    progress_bar.UpdateBar(28, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet = workbook["Егорьевск Copy"]
    ws3 = worksheet.title = "Шатура"
    progress_bar.UpdateBar(29, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    FullRange = "C6:" + get_column_letter(worksheet.max_column) \
                + str(worksheet.max_row)
    progress_bar.UpdateBar(30, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet.auto_filter.ref = FullRange
    progress_bar.UpdateBar(31, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    ws3 = worksheet['D4'] = shatoor1
    ws3 = worksheet['E4'] = shatoor2
    ws3 = worksheet['G4'] = shatoor3
    ws3 = worksheet['C4'] = 'Шатура'
    ws3 = worksheet.auto_filter.add_filter_column(0, shatoora)
    progress_bar.UpdateBar(32, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    safepath = today.strftime("%d,%m,%Y" + ".xlsx")
    save_filename = value["folder_s"] + "/map4_" + tosafe[0] + "_" + tosafe[
        1] + ".xlsx"  # + re.findall(r'\d{2}.\d{2}.\d{2}', dataSheet.value)

    if os.path.isfile(save_filename):
        expand = 1
        while True:
            expand += 1
            new_file_name = save_filename.split(".xlsx")[0] + " (" + str(expand) + ")" + ".xlsx"
            if os.path.isfile(new_file_name):
                continue
            else:
                save_filename = new_file_name
                break
    try:
        # save_filename = path

        # path2 =
        # sg.popup(f"Saved: {save_filename}")
        # path2 = easygui.filesavebox(default="HOME/APPDATA/RPBeta/Egoryevsk.xlsx", filetypes='*.xslx')
        workbook.save(tmp_file)
        shutil.copy(tmp_file, save_filename)
    except Exception as e:
        try:
            workbook.save(tmp_file)
            shutil.copy(tmp_file, save_filename)
        except Exception as e:
            sg.popup_auto_close(e, 'Вы отказались от сохранения. \nНо файл сохранился в той же папке.')
            # sg.popup_ok(e)
            workbook.save(path + safepath)
            progress_bar.UpdateBar(0, 20)
        return (safepath)
    progress_bar.UpdateBar(33, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    sg.popup_ok('Файл готов!')
    progress_bar.UpdateBar(0, 33)



def button98():
    progress_bar.UpdateBar(0, 3)  ##############PROGRESS
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    size = (50, 3)
    auto_size_button = True
    path = easygui.fileopenbox(default="HOME/Downloads/*.xls", filetypes='*.xls')
    fname = path

    progress_bar.UpdateBar(1, 3)  ##############PROGRESS
    try:
        wb = excel.Workbooks.Open(fname)
        wb.SaveAs(fname + "x", FileFormat=51)  # FileFormat = 51 is for .xlsx extension
        progress_bar.UpdateBar(2, 3)  ##############PROGRESS
        wb.Close()

        progress_bar.UpdateBar(3, 3)  ##############PROGRESS
        sg.popup_auto_close('сконвертировано')
    except Exception as e:
        try:
            wb = excel.Workbooks.Open(fname)
        except Exception as e:
            sg.popup_auto_close(e)
            #sg.popup_auto_close('неудача')
            wb = None
            progress_bar.UpdateBar(0, 3)
            return (wb)
    finally:
        #sg.popup_auto_close('завершение')  # задача при которой условия частично выполнены.
        excel.Application.Quit()
        progress_bar.UpdateBar(0, 3)


def about_me():
    #sg.PopupQuick('"Все великое начинается с малого." - Peter Senge', auto_close=False)

    my_text = "\nПрограмма написана саня).\n" \
              "\nДата последнего редактирования: 20.07.2021" \
                  "\nВерсия программы: " + versionRR + \
              "\nПрограмма и исходный код\n" \
              "распространяются по лицензии" \
              "\nGNU General Public License v3.0\n" \
    "\n" \
              "\n" \
              "\n"

    sg.popup('О программе', 'Добро пожаловать в программу РНИСка Отчеты!', my_text)

def howto():
    #sg.PopupQuick('"Все великое начинается с малого." - Peter Senge', auto_close=False)

    my_text = "\nДля успешной работы необходимо сконвертировать отчет из рнис" \
              '\nВ РНИСе он должен называться как "Итоговый отчет о работе выходов по филиалу (общие показатели)"' \
              '\nПосле выбора нужного файла произойдет конвертация!\nОригинальный файл никак не изменяется.' \
              '\nВыберете место сохранения и дождитесь окончания процесса.' \
              '\nНажмите на нужную кнопку какого парка вы хотите получить в готовом отчете и выбирайте файл который' \
              ' сохраняли при конвертации.' \
              '\nЗатем выберете место сохранения файла и по окончанию процесса, перейдите в папку сохранения отчета.' \
              '\nИ и на этом все!' \
              '\nПожалуйста не забывайте что после создания файла с отчетом, нужно "принять" сортировку и нажать на' \
              ' стрелочку в поле "РЕГ НОМЕР" и далее на "ОК". По всем вопросам можете обращаться ко мне.' \
              '\nФайл отчетов можете как до так и после конвертации произвольно называть и изменять.'
    sg.popup('Инструкция для программы', 'Добро пожаловать в программу РНИСка Отчеты!', my_text)

def button99():
    imwatchingyou.show_debugger_window()


# кнопочки для проверки белого списка
dispatch_dictionary = { 'Егорьевск':button1, 'Раменское':button2, 'Шатура':button3, 'МАП4':button4, 'Конверт':button98,
                       'О программе':about_me, 'Инструкция':howto, 'debugHigh':button99 }

menu_layout: list = [['RNISka Reports', ['Выход']],
                     ['Окно', ['О программе', 'debugHigh', 'debugLight']],
                     ['Справка', ['Инструкция']]]
# кнопки для конкретно гуи

#sg.Window.get_screen_size()
#w, h = sg.Window.get_screen_size()
#getResW = int(w / 2 + (w / 20))
#getResH = int((w + h) / 4)
#buttonRes = int()
#WIN_H = int(getResH / 714 + 1)
#WIN_W = int(getResW / 44 + 1)
#textSi = int((getResW * getResH) / 49152)
textSi = 12
WIN_W: int = 12
WIN_H: int = 3
defDownFold = os.path.expanduser(r"~\Downloads")
getItdef = defDownFold.replace("\\", "/")
layout: list = [[sg.Menu(menu_layout)],
                [sg.ProgressBar(1, orientation='h', size=(120, 20), key='progress', pad=((1, 1), 1))],
                [sg.Text('Добро пожаловать!', relief='sunken', auto_size_text=True, justification='center',
                         font=('Consolas', textSi), size=(200, 1))],
                [sg.Text('Сверху есть кнопка "О программе" а внутри нее есть "Инфо"!', auto_size_text=True,
                         justification='center', font=('Consolas', textSi), size=(200, 1))],
                [sg.Text('Нажмите на нее чтобы прочитать инструкцию.', auto_size_text=True, justification='center',
                         font=('Consolas', textSi), size=(200, 1))],

                    [sg.T('Место сохранения')],
                    [sg.In(getItdef, key='folder_s'), sg.FolderBrowse("Обзор", target='folder_s')],

                [sg.Button('Егорьевск', font=('Consolas', textSi), auto_size_button=True, tooltip=(
                    'Создается один лист с Егорьевском.\nУбедитесь что вы выбрали файл "Итоговый отчет о работе выходов по филиалу (общие показатели)"'),
                           size=(WIN_W, WIN_H), border_width=(5), pad=((1, 1), 1)),
                 sg.Button('Раменское', font=('Consolas', textSi), auto_size_button=True, tooltip=(
                     'Создается один лист с Раменское.\nУбедитесь что вы выбрали файл "Итоговый отчет о работе выходов по филиалу (общие показатели)"'),
                           size=(WIN_W, WIN_H), border_width=(5), pad=((1, 1), 1)),
                 sg.Button('Шатура', font=('Consolas', textSi), auto_size_button=True, tooltip=(
                     'Создается один лист с Шатурой.\nУбедитесь что вы выбрали файл "Итоговый отчет о работе выходов по филиалу (общие показатели)"'),
                           size=(WIN_W, WIN_H), border_width=(5), pad=((1, 1), 1))],
                [sg.Button('МАП4', font=('Consolas', textSi), auto_size_button=True, tooltip=(
                    'Создается 3 листа всего МАПа.\nУбедитесь что вы выбрали файл "Итоговый отчет о работе выходов по филиалу (общие показатели)"'),
                           size=(WIN_W, WIN_H), border_width=(5), pad=((1, 1), 1))],
                [sg.Button('Конверт', font=('Consolas', textSi), auto_size_button=True, size=(WIN_W, WIN_H),
                           border_width=(5), pad=((1, 1), 1))],
                [sg.Quit('Выход', font=('Consolas', textSi), auto_size_button=True, size=(35, 3), pad=((1, 1), 100))]]

# титульное окно
window = sg.Window(testWindows, layout, size=(1000, 720), icon=r"img/icon.ico", element_justification='c')
progress_bar = window.FindElement('progress')

# белый список
while True:
    # ифелс
    event, value = window.read()
    #print(event, value)
    #if event in ('Выход', sg.WIN_CLOSED):
     #   window.close(); del window
    if event in (about_me, 'n:78'):
        about_me()

    #if event == 'FolderBrowse':
     #   foldername = sg.PopupGetFolder('Select folder', no_window=True)



    elif event in ('Выход', sg.WIN_CLOSED):  # Window close button event
        break
    #imwatchingyou.refresh_debugger()

    # белый список в действии

    if event in dispatch_dictionary:
        func_to_call = dispatch_dictionary[event]   # словарь в действии
        func_to_call()
    else:
        sg.Print('функции "{}" не существует в данной версии'.format(event))

window.close()


#window.close()
#del window

